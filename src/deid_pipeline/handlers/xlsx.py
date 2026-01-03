from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Optional

from .base import ExtractedDocument, SegmentSpec, build_document_from_segments
from ..core.contracts import replacement_key


class XlsxHandler:
    extensions = [".xlsx"]

    def extract(self, input_path: Path, *, language: str) -> ExtractedDocument:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as exc:
            raise ImportError("openpyxl is required to extract XLSX files") from exc

        wb = load_workbook(filename=str(input_path), read_only=True, data_only=True)
        segments: list[SegmentSpec] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    text = str(cell.value)
                    if not text:
                        continue
                    segments.append(
                        SegmentSpec(
                            text=text,
                            cell={
                                "sheet": sheet_name,
                                "address": cell.coordinate,
                                "row": int(cell.row),
                                "col": int(cell.column),
                            },
                        )
                    )

        wb.close()

        return build_document_from_segments(
            input_path=input_path,
            language=language,
            segments=segments or [SegmentSpec(text="")],
            separator="\n",
            metadata={"format": "xlsx", "sheets": wb.sheetnames},
        )

    def rebuild(
        self,
        document: ExtractedDocument,
        *,
        output_text: str,
        entities: list[dict],
        replacement_map: Dict[str, str],
        events: list[dict],
        output_dir: Optional[Path] = None,
        mode: str = "replace",
    ) -> Dict[str, Any]:
        artifacts: Dict[str, Any] = {"output_text": output_text}

        if output_dir is None or mode != "replace":
            artifacts["rebuild_supported"] = False
            return artifacts

        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            artifacts["rebuild_supported"] = False
            return artifacts

        output_dir.mkdir(parents=True, exist_ok=True)
        out_path = output_dir / f"{document.input_path.stem}.deid.xlsx"

        wb = load_workbook(filename=str(document.input_path))

        for entity in entities:
            cell_ref = entity.get("cell")
            if not cell_ref:
                continue

            sheet = cell_ref.get("sheet")
            address = cell_ref.get("address")
            if not sheet or not address:
                continue
            if sheet not in wb.sheetnames:
                continue

            original = entity.get("text")
            entity_type = entity.get("type")
            if not original or not entity_type:
                continue

            repl = replacement_map.get(replacement_key(str(entity_type), str(original)))
            if repl is None:
                continue

            ws = wb[sheet]
            cell = ws[address]
            value = "" if cell.value is None else str(cell.value)
            cell.value = value.replace(str(original), str(repl), 1)

        wb.save(str(out_path))
        wb.close()

        artifacts["output_path"] = str(out_path)
        artifacts["rebuild_supported"] = True
        artifacts["rebuild_mode"] = "xlsx_cell_replace"
        return artifacts
